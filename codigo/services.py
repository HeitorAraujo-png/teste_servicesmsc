import os
from settings import *
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class Relatorio:

    def __init__(self, arquivos):
        """Adicione o relatorio de 1 a 15 dias e de 15 a 31"""
        self.LinkPathPTD = rf"{BASE_DIR}{UPLOADCSV}\DPTDIA.csv"
        self.Concatena(arquivos)
        self.dias = self.Dias()
        self.departamento = pd.read_csv(self.LinkPathPTD, encoding="latin1")
        self.nomes, self.centros = self.CentroCustos()
        self.Adicionar()
    
    def Concatena(self, arquivos):
        """Concatena os dois arquivos .XLSX tranformando em um unico arquivo .CSV"""
        self.departamento = pd.DataFrame()
        for i in arquivos:
            i = fr'{BASE_DIR}\{RELATORIO}\{i}'
            relatorio = pd.read_excel(i, parse_dates=["DATA"])
            relatorio["DATA"] = relatorio["DATA"].dt.strftime("%d/%m/%y")
            self.departamento = pd.concat([self.departamento, relatorio], ignore_index=True)
        self.departamento.to_csv(self.LinkPathPTD, encoding="latin1")

    def Adicionar(self):
        """Adiciona os dias no arquivo .CSV que contem as principais informações"""

        colunas = ['DPT'] + self.dias + ['Total', 'Valor total', '%']
        lista = {col: None for col in colunas}
        self.relatorio = pd.DataFrame([lista])
        self.relatorio = self.relatorio.drop(0)
        self.relatorio.to_csv(fr'{BASE_DIR}\{UPLOADCSV}\relatorio.csv')

    def remove(self, lista):
        """Remove itens duplicados de uma lista"""
        return list(dict.fromkeys(sorted(lista)))

    def Dias(self):
        """Volta uma lista com todos os dias"""
        return self.remove(list(self.departamento["DATA"]))

    def TakeIndex(self, search):
        """Pega index de coluna especifica\n
    Exemplo: .to_list()) vai voltar uma lista com todas as colunas do arquivo
    ['dpt', '28/11/2024', '29/11/2024', '30/11/2024', 'Total', 'Valor total', '%']
    Vamos dizer que search = 'Valor total'
    index = self.relatorio.columns.to_list().index(search vai retornar 5
    return result == 'F'
        """
        index = self.relatorio.columns.to_list().index(search)
        result = ""
        while index >= 0:
            result = chr(index % 26 + 65) + result
            index = index // 26 - 1
        return result

    def CentroCustos(self):
        """Classifica por ordem numerica cada centro e custo e adiciona os nomes dos mesmos"""
        centro = self.departamento["C.C"]
        centro = self.remove(list(centro))
        lista = []
        for i in centro:
            i = str(i)
            ic = f'{i[0]}{i[1]}'
            cc = f'{i[2]}{i[3]}'
            if ic == "81":
                emp = "TVCA"
            if ic == "65":
                emp = "Portal MT"
            if ic == "69":
                emp = "FMCA"
            if ic == "85":
                emp = "On Line(MT)"
            if cc == "03":
                tipo = "ADM"
            if cc == "06":
                tipo = "RH"
            if cc == "10":
                tipo = "COMERCIA"
            if cc == "11":
                tipo = "OPEC"
            if cc == "7":
                tipo = "MKT"
            if cc == "14":
                tipo = "PROGRAMAÇÃO"
            if cc == "15":
                tipo = "JORNALISMO"
            if cc == "21":
                tipo = "TECNOLOGIA"
            lista.append(f"{i} - {tipo} - {emp}")
        return lista, centro

    def Converte(self):
        """Converte o arquivo .CSV em .XLSX"""
        coluna = 2
        atual = {"DPT": 0}
        for dia in self.dias:
            total_dia = 0
            for cc in self.centros:
                qtd = len(
                    self.departamento[
                        (self.departamento["C.C"] == cc)
                        & (self.departamento["DATA"] == dia)
                    ]
                )
                total_dia += qtd
            atual[dia] = (
                f"=SUM({self.TakeIndex(dia)}{coluna}:{self.TakeIndex(dia)}{len(self.centros) + 1})"
            )
        atual["Valor total"] = (
            f"=SUM({self.TakeIndex('Valor total')}{coluna}:{self.TakeIndex('Valor total')}{len(self.centros) + 1})"
        )
        atual["Total"] = (
            f"=SUM({self.TakeIndex('Total')}{coluna}:{self.TakeIndex('Total')}{len(self.centros) + 1})"
        )
        for cc, nome in zip(self.centros, self.nomes):
            nova_linha = {"DPT": nome}
            soma = 0
            for dia in self.dias:
                qtd = len(
                    self.departamento[
                        (self.departamento["C.C"] == cc)
                        & (self.departamento["DATA"] == dia)
                    ]
                )
                nova_linha[dia] = qtd
                soma += qtd
            nova_linha["%"] = (
                f"=SUM({self.TakeIndex('Valor total')}{coluna}/{self.TakeIndex('Valor total')}{len(self.centros) + 2})"
            )
            nova_linha["Total"] = (
                f"=SUM({self.TakeIndex(self.dias[0])}{coluna}:{self.TakeIndex(self.dias[-1])}{coluna})"
            )
            nova_linha["Valor total"] = f"=${self.TakeIndex('Total')}${coluna} * 20"
            self.relatorio = pd.concat(
                [self.relatorio, pd.DataFrame([nova_linha])], ignore_index=True
            )
            coluna = coluna + 1
        atual["%"] = (
            f"=SUM({self.TakeIndex('%')}{2}:{self.TakeIndex('%')}{len(self.centros) + 1})"
        )
        self.relatorio = pd.concat(
            [self.relatorio, pd.DataFrame([atual])], ignore_index=True
        )
        self.relatorio.to_excel(
                rf"{BASE_DIR}\{UPLOADXLSX}\relatorio_atualizado.xlsx", index=False)

    def Espaco(self):
        """Arruma o espaçamento de cada coluna do arquivo em excel"""
        self.wb = load_workbook(rf"{BASE_DIR}\{UPLOADXLSX}\relatorio_atualizado.xlsx")
        sheet = self.wb.sheetnames[0]
        ws = self.wb[sheet]
        for col in ws.columns:
            max_l = 0
            coluna = col[0].column
            coluna_letra = get_column_letter(coluna)
            for cell in col:
                try:
                    if cell.value:
                        max_l = max(max_l, len(str(cell.value)))
                except Exception as e:
                    pass
            ajuste = max_l + 2
            ws.column_dimensions[coluna_letra].width = ajuste
            self.Salva()
        
    def Salva(self):
        hoje = datetime.now().strftime("%d-%m-%y-")
        if os.path.exists((fr"{BASE_DIR}\{UPLOADXLSX}\relatorio_atualizado.xlsx")):
            self.wb.save(fr"{BASE_DIR}\{HISTORICO}\{hoje}relatorio_atualizado.xlsx")
            self.relatorio.to_excel(
                rf"{BASE_DIR}\{HISTORICO}\{hoje}relatorio_atualizado.xlsx", index=False)
        else:
            self.wb.save(fr"{BASE_DIR}\{UPLOADXLSX}\relatorio_atualizado.xlsx")
