from openpyxl import load_workbook
from datetime import datetime
from shutil import move
import pandas as pd
from path import *
from db import *
import calendar
import datetime
import logging
import os


class Querys:

    def __init__(self):
        self.util = Utils
        self.POG()
        
    
    def Concatena(arquivos):
        """Concatena os dois arquivos .XLSX tranformando em um unico arquivo .CSV"""
        departamento = pd.DataFrame()
        relatorio = pd.read_excel(arquivos, parse_dates=["DATA"])
        relatorio["DATA"] = relatorio["DATA"].dt.strftime("%d/%m/%y")
        departamento = pd.concat([departamento, relatorio], ignore_index=True)
        departamento.to_csv(LINK, encoding="latin1")

    def Create(self, CenCus, cd, day, hora):
        # existe = session.query(Pessoa).filter_by(horario=hora,cc=CenCus,dia=day,cod=cd).first()
        # if existe:
        #     # nome
        #     self.util.Log(f'Usuario já cadastrado! {cd}')
        # else:
        pessoa = Pessoa(cc=CenCus, cod=cd, dia=day, horario=hora)
        session.add(pessoa)
        session.commit()


    def Centro_Custos(self):
        lista = []
        query = session.query(Pessoa).order_by("cc")
        for i in query:
            lista.append(i.cc)
        centros = self.util.remove(lista)
        return centros

    def Filtro(self, dia, cc):
        num = session.query(Pessoa).filter_by(dia=dia, cc=cc)
        return num.count()

    def Nomeclatura(self, cc):
        csv = pd.read_csv(fr"{BASE_DIR}\nomeclatura.csv", sep="|", encoding="latin1")
        for dpt, nome in zip(csv.DPT, csv.nome):
            if dpt == cc:
                return f"{dpt} - {nome}"

    def NomeMes(self, data):
        csv = pd.read_csv(fr"{BASE_DIR}\nomeclatura.csv", sep="|", encoding="latin1")
        for mes, nome in zip(csv.mes, csv.NomeMes):
            mes = int(mes) if mes > 9 else f'0{int(mes)}'
            if mes == data:
                
                return nome
        self.util.Log("Erro ao criar arquivo ")

    def MakeCSV(self):
        self.POG()
        arquivo = pd.DataFrame()
        nome = f"Relatorio-{self.NomeMes(self.mes_atual)}"
        nome_anterior =  f"Relatorio-{self.NomeMes(self.mes_anterior)}"
        self.path = fr"{BASE_DIR}{UPLOADCSV}\{nome}.csv"
        self.xlsx = fr"{BASE_DIR}{UPLOADXLSX}\{nome}.xlsx"
        xlsx_anterior = fr"{BASE_DIR}{UPLOADXLSX}\{nome_anterior}.xlsx"
        arquivo.to_csv(self.path, index=False, encoding='latin1')
        return self.path, self.xlsx, xlsx_anterior, nome_anterior

        
    def TakeIndex(self, search):
        csv = pd.read_csv(self.path)
        index = csv.columns.to_list().index(search)
        result = ""
        while index >= 0:
            result = chr(index % 26 + 65) + result
            index = index // 26 - 1
        return result
    
    def Linhas(self):
        colunas = ["DPT"] + self.dias + ["Total", "Valor total", "%"]
        lista = {col: None for col in colunas}
        arquivo = pd.DataFrame([lista])
        arquivo = arquivo.drop(0)
        arquivo.to_csv(self.path, index=False, encoding='latin1')
        return arquivo.columns.to_list()
    
    def Espaco(self):
        """Arruma o espaçamento de cada coluna do arquivo em excel"""
        self.wb = load_workbook(self.xlsx)
        ws = self.wb[self.wb.sheetnames[0]]
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions[self.TakeIndex('Valor total')].width = 12
        self.Salva()
        
    def Salva(self):
        # hoje = datetime.now().strftime("%d-%m-%y-")
        self.wb.save(self.xlsx)

    def POG(self):
        self.mes_atual = int(datetime.datetime.now().strftime('%m'))
        self.ano = int(datetime.datetime.now().strftime('%y'))
        if int(datetime.datetime.now().strftime('%m')) > 0:
            self.mes_anterior =  int(datetime.datetime.now().strftime('%m')) - 1 
        else:
            self.mes_anterior = 12
            self.ano -= 1
        max_dia_atual = calendar.monthrange(self.ano, self.mes_atual)[1]
        max_dia_anterior = calendar.monthrange(self.ano, self.mes_anterior)[1]
        self.mes_atual = self.mes_atual if self.mes_atual> 9 else f'0{self.mes_atual}'
        self.mes_anterior = self.mes_anterior if self.mes_anterior> 9 else f'0{self.mes_anterior}'
        self.dias = ([f'{dia}/{self.mes_atual}/{self.ano}' if dia > 9 else f'0{dia}/{self.mes_atual}/{self.ano}' for dia  in range(1, max_dia_anterior + 1)])
        self.dias_anterior = ([f'{dia}/{self.mes_anterior}/{self.ano}' if dia > 9 else f'0{dia}/{self.mes_anterior}/{self.ano}' for dia  in range(1, max_dia_atual + 1)])
        return self.dias, self.dias_anterior

    def Drop(self):
        sql = pd.read_sql('SELECT * FROM Pessoas', session.bind)
        sql.to_csv(fr'{BASE_DIR}{UPLOADCSV}\Db-{datetime.now().strftime('%d-%m-%Y')}.csv')
        session.query(Pessoa).delete()
        session.commit()

class Utils:
        
    def __init__(self):
        pass
        
    def Log(logmsg):
        logging.basicConfig(filemode='a',
        filename=f'{BASE_DIR}{DEV_LOG}', 
        level=logging.INFO,
        format="%(asctime)s | %(process)d | %(message)s",
        datefmt="%d-%m-%y %H:%M:%S",
        )
        logging.info(logmsg)
        
    def Converter(dia_str, hora_str):
        try:
            dia = datetime.strptime(dia_str, '%d/%m/%Y').date()
            horario = datetime.strptime(hora_str, '%H:%M')
            return dia, horario
        except Exception as ex:
            Utils.Log(f'Erro ao criar usuario: {ex}')

    def remove(lista):
        return list(dict.fromkeys(sorted(lista)))